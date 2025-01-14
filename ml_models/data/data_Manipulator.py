import openpyxl

file_path_1 = 'raw_results/original_student_data.xlsx'

# Load the Excel file
workbook = openpyxl.load_workbook(file_path_1)
worksheet = workbook.active

for col in range(1, worksheet.max_column + 1):
    current_header = worksheet.cell(row=1, column=col).value
    if current_header:  # Check if header exists
        worksheet.cell(row=1, column=col, value=current_header.lower())

name_col_idx = None
biology_col_idx = None
biology_blind_col_idx = None

# Find the column index for the 'Name', 'Biology', and 'Biology (Blind)' columns
for col in range(1, worksheet.max_column + 1):
    header = worksheet.cell(row=1, column=col).value
    if header == 'name':
        name_col_idx = col
    elif header == 'biology':
        biology_col_idx = col
    elif header == 'biology for the blind':
        biology_blind_col_idx = col

# Merging the data from the two columns
if biology_col_idx and biology_blind_col_idx:
    for row in range (2, worksheet.max_row + 1):
        biology_blind_value = worksheet.cell(row=row, column=biology_blind_col_idx).value
        if biology_blind_value:
            worksheet.cell(row=row, column=biology_col_idx, value=biology_blind_value)

# Iterate through the rows and modify the Student ID
for row in range(2, worksheet.max_row + 1):  # Start from row 2 (assuming header row is in row 1)
    # Modify the Student ID with the custom format "SID-01" onwards
    worksheet.cell(row=row, column=1, value=f"SID-{row-1:02}")

# Delete the 'Biology (Blind)' and 'Name' columns
if biology_blind_col_idx:
    worksheet.delete_cols(biology_blind_col_idx)
if name_col_idx:
    worksheet.delete_cols(name_col_idx)

worksheet.cell(row=1, column=1, value="student_id")

# Save the modified Excel file
workbook.save('data_results/student_data.xlsx')